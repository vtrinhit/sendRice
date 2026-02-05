"""
Excel Parser Service
Parses Excel files to extract employee salary data.
"""
import os
from typing import List, Dict, Any, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.cell import Cell
import uuid
from datetime import datetime

from app.config import settings


class ExcelParserService:
    """Service for parsing Excel files."""

    def __init__(self, file_path: str):
        """Initialize with Excel file path."""
        self.file_path = file_path
        self.workbook = None

    def __enter__(self):
        """Context manager entry - load workbook."""
        self.workbook = load_workbook(self.file_path, data_only=True)
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit - close workbook."""
        if self.workbook:
            self.workbook.close()

    def get_sheet_names(self) -> List[str]:
        """Get list of sheet names in the workbook."""
        if not self.workbook:
            raise ValueError("Workbook not loaded. Use context manager.")
        return self.workbook.sheetnames

    def get_sheet(self, sheet_name: str) -> Worksheet:
        """Get a specific worksheet."""
        if not self.workbook:
            raise ValueError("Workbook not loaded. Use context manager.")
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
        return self.workbook[sheet_name]

    def parse_employees(
        self,
        sheet_name: str,
        header_row: int = 1,
        data_start_row: int = 2,
        code_column: str = "A",
        name_column: str = "B",
        phone_column: str = "C",
        salary_column: str = "D",
    ) -> List[Dict[str, Any]]:
        """
        Parse employee data from a worksheet.

        Args:
            sheet_name: Name of the sheet to parse
            header_row: Row number containing headers (1-indexed)
            data_start_row: Row number where data starts (1-indexed)
            code_column: Column letter for employee code
            name_column: Column letter for employee name
            phone_column: Column letter for phone number
            salary_column: Column letter for salary

        Returns:
            List of employee dictionaries
        """
        sheet = self.get_sheet(sheet_name)
        employees = []

        # Get column indices
        code_col_idx = column_index_from_string(code_column.upper())
        name_col_idx = column_index_from_string(name_column.upper())
        phone_col_idx = column_index_from_string(phone_column.upper())
        salary_col_idx = column_index_from_string(salary_column.upper())

        # Parse data rows
        row_num = data_start_row
        while True:
            name_cell = sheet.cell(row=row_num, column=name_col_idx)

            # Stop if name is empty (end of data)
            if not name_cell.value or str(name_cell.value).strip() == "":
                break

            code_cell = sheet.cell(row=row_num, column=code_col_idx)
            phone_cell = sheet.cell(row=row_num, column=phone_col_idx)
            salary_cell = sheet.cell(row=row_num, column=salary_col_idx)

            employee = {
                "row_number": row_num,
                "employee_code": self._get_cell_value_str(code_cell),
                "name": str(name_cell.value).strip(),
                "phone": self._normalize_phone(phone_cell.value),
                "salary": self._parse_salary(salary_cell.value),
            }
            employees.append(employee)
            row_num += 1

        return employees

    def get_employee_row_data(
        self,
        sheet_name: str,
        row_number: int,
        start_col: str = "A",
        end_col: str = "G"
    ) -> List[List[Any]]:
        """
        Get data for a specific employee row range for image generation.

        Returns cell data including values and basic formatting info.
        """
        sheet = self.get_sheet(sheet_name)
        start_col_idx = column_index_from_string(start_col.upper())
        end_col_idx = column_index_from_string(end_col.upper())

        row_data = []
        for col_idx in range(start_col_idx, end_col_idx + 1):
            cell = sheet.cell(row=row_number, column=col_idx)
            cell_data = {
                "value": self._format_cell_value(cell),
                "column": get_column_letter(col_idx),
                "row": row_number,
            }
            row_data.append(cell_data)

        return row_data

    def get_salary_image_data(
        self,
        sheet_name: str,
        employee_row: int,
        start_row: Optional[int] = None,
        end_row: Optional[int] = None,
        start_col: str = "A",
        end_col: str = "G"
    ) -> Dict[str, Any]:
        """
        Get data for generating salary image.

        If start_row/end_row not specified, gets a range around the employee row.
        """
        sheet = self.get_sheet(sheet_name)

        # Default to showing rows around the employee
        if start_row is None:
            start_row = max(1, employee_row - 2)
        if end_row is None:
            end_row = employee_row + 2

        start_col_idx = column_index_from_string(start_col.upper())
        end_col_idx = column_index_from_string(end_col.upper())

        # Get column widths
        col_widths = {}
        for col_idx in range(start_col_idx, end_col_idx + 1):
            col_letter = get_column_letter(col_idx)
            col_dim = sheet.column_dimensions.get(col_letter)
            col_widths[col_letter] = col_dim.width if col_dim and col_dim.width else 10

        # Get all cell data
        rows_data = []
        for row_num in range(start_row, end_row + 1):
            row_data = []
            for col_idx in range(start_col_idx, end_col_idx + 1):
                cell = sheet.cell(row=row_num, column=col_idx)
                cell_info = {
                    "value": self._format_cell_value(cell),
                    "column": get_column_letter(col_idx),
                    "row": row_num,
                    "is_employee_row": row_num == employee_row,
                    "merged": self._is_merged(sheet, row_num, col_idx),
                }
                row_data.append(cell_info)
            rows_data.append(row_data)

        return {
            "rows": rows_data,
            "col_widths": col_widths,
            "start_row": start_row,
            "end_row": end_row,
            "start_col": start_col,
            "end_col": end_col,
            "employee_row": employee_row,
        }

    def _get_cell_value_str(self, cell: Cell) -> Optional[str]:
        """Get cell value as string."""
        if cell.value is None:
            return None
        return str(cell.value).strip()

    def _normalize_phone(self, value: Any) -> Optional[str]:
        """Normalize phone number to string format."""
        if value is None:
            return None

        # Convert to string and clean
        phone = str(value).strip()

        # Remove common prefixes and formatting
        phone = phone.replace(" ", "").replace("-", "").replace(".", "")

        # Handle float representation (e.g., 901234567.0)
        if "." in phone:
            phone = phone.split(".")[0]

        # Add leading zero if missing (Vietnamese numbers)
        if phone and not phone.startswith("0") and len(phone) == 9:
            phone = "0" + phone

        return phone if phone else None

    def _parse_salary(self, value: Any) -> Optional[int]:
        """Parse salary value to integer."""
        if value is None:
            return None

        if isinstance(value, (int, float)):
            return int(value)

        # Try parsing string
        try:
            # Remove common formatting
            cleaned = str(value).replace(",", "").replace(".", "").replace(" ", "")
            cleaned = cleaned.replace("VND", "").replace("đ", "").strip()
            return int(cleaned) if cleaned else None
        except (ValueError, TypeError):
            return None

    def _format_cell_value(self, cell: Cell) -> str:
        """Format cell value for display."""
        if cell.value is None:
            return ""

        value = cell.value

        # Format numbers with thousand separators
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            if isinstance(value, float) and value.is_integer():
                return f"{int(value):,}".replace(",", ".")
            elif isinstance(value, int):
                return f"{value:,}".replace(",", ".")
            else:
                return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

        return str(value)

    def _is_merged(self, sheet: Worksheet, row: int, col: int) -> bool:
        """Check if a cell is part of a merged range."""
        for merged_range in sheet.merged_cells.ranges:
            if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= col <= merged_range.max_col):
                return True
        return False


async def parse_excel_file(
    file_path: str,
    sheet_name: str,
    config: Dict[str, Any]
) -> Tuple[List[Dict[str, Any]], str]:
    """
    Async wrapper for parsing Excel file.

    Returns:
        Tuple of (employees list, actual sheet name used)
    """
    with ExcelParserService(file_path) as parser:
        # If sheet name not specified, use first sheet
        if not sheet_name:
            sheets = parser.get_sheet_names()
            sheet_name = sheets[0] if sheets else "Sheet1"

        employees = parser.parse_employees(
            sheet_name=sheet_name,
            header_row=config.get("header_row", 1),
            data_start_row=config.get("data_start_row", 2),
            code_column=config.get("code_column", "A"),
            name_column=config.get("name_column", "B"),
            phone_column=config.get("phone_column", "C"),
            salary_column=config.get("salary_column", "D"),
        )

        return employees, sheet_name
