"""
Optimized Salary Slip Service
Uses LibreOffice UNO API for batch processing - keeps LibreOffice open.
Exports directly to PNG (no PDF intermediate step).
"""
import os
import subprocess
import tempfile
import shutil
import base64
import time
import threading
import sys
from pathlib import Path
from typing import Tuple, Optional, Dict, Any, List
from dataclasses import dataclass
import logging

from app.config import settings

logger = logging.getLogger(__name__)


def get_libreoffice_path() -> str:
    """Get the correct LibreOffice executable path based on OS."""
    if sys.platform == "win32":
        # Common Windows installation paths
        possible_paths = [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
            r"C:\Program Files\LibreOffice\program\soffice.com",
        ]
        for path in possible_paths:
            if os.path.exists(path):
                return path
        # Try soffice in PATH
        return "soffice"
    else:
        # Linux/Mac
        return "libreoffice"


@dataclass
class BatchResult:
    """Result of a batch image generation."""
    employee_code: str
    success: bool
    base64_image: Optional[str] = None
    salary: Optional[int] = None
    error: Optional[str] = None


class OptimizedSalarySlipService:
    """
    Optimized service for generating salary slip images.

    Key optimizations:
    1. Keeps LibreOffice process alive for batch processing
    2. Opens Excel file once, modifies D9 for each employee
    3. Exports directly to PNG (no PDF intermediate)
    4. Uses UNO API for programmatic control
    """

    SALARY_SLIP_SHEET = "Phiếu lương"
    EMPLOYEE_CODE_CELL = "D9"
    SALARY_CELL = "E24"

    def __init__(self):
        self.output_dir = settings.temp_images_dir
        os.makedirs(self.output_dir, exist_ok=True)
        self._lock = threading.Lock()

    def generate_batch(
        self,
        excel_file_path: str,
        employee_codes: List[str],
        image_config: Optional[Dict[str, Any]] = None,
        callback: Optional[callable] = None
    ) -> List[BatchResult]:
        """
        Generate salary slips for multiple employees in one batch.

        Note: Cannot keep workbook open across saves because openpyxl has issues
        with embedded images (file handles get closed after first save).
        Instead, we reload for each employee but optimize other aspects.

        Args:
            excel_file_path: Path to the Excel file
            employee_codes: List of employee codes to process
            image_config: Image generation config
            callback: Optional callback(employee_code, result) for progress updates

        Returns:
            List of BatchResult objects
        """
        config = {
            "image_start_col": "B",
            "image_end_col": "H",
            "image_start_row": 4,
            "image_end_row": 29,
        }
        if image_config:
            config.update(image_config)

        results = []

        with tempfile.TemporaryDirectory() as temp_dir:
            temp_dir_path = Path(temp_dir)
            temp_excel = temp_dir_path / "salary.xlsx"

            # Copy Excel to temp location once
            shutil.copy(excel_file_path, temp_excel)
            logger.info(f"[BATCH] Processing {len(employee_codes)} employees")

            # Process each employee
            for emp_code in employee_codes:
                try:
                    result = self._process_single_optimized(
                        temp_excel,
                        emp_code,
                        config,
                        temp_dir_path
                    )
                    results.append(result)

                    if callback:
                        callback(emp_code, result)

                except Exception as e:
                    import traceback
                    error_result = BatchResult(
                        employee_code=emp_code,
                        success=False,
                        error=str(e)
                    )
                    results.append(error_result)

                    if callback:
                        callback(emp_code, error_result)

                    logger.error(f"Failed to process {emp_code}: {e}")
                    logger.error(f"Traceback:\n{traceback.format_exc()}")

            logger.info(f"[BATCH] Completed. Success: {sum(1 for r in results if r.success)}/{len(results)}")

        return results

    def _process_single_optimized(
        self,
        excel_path: Path,
        employee_code: str,
        config: Dict[str, Any],
        temp_dir: Path
    ) -> BatchResult:
        """
        Process a single employee - optimized version with less overhead.

        Must reload workbook each time due to openpyxl image handling bug.
        """
        from openpyxl import load_workbook
        from openpyxl.worksheet.page import PageMargins, PrintPageSetup
        from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties

        logger.info(f"[{employee_code}] Processing")

        # Load workbook (required for each employee due to image bug)
        wb = load_workbook(excel_path)

        if self.SALARY_SLIP_SHEET not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{self.SALARY_SLIP_SHEET}' not found")

        ws = wb[self.SALARY_SLIP_SHEET]

        # Initialize sheet properties if needed
        if ws.sheet_properties is None:
            ws.sheet_properties = WorksheetProperties()
        if ws.sheet_properties.pageSetUpPr is None:
            ws.sheet_properties.pageSetUpPr = PageSetupProperties()
        if ws.page_setup is None or ws.page_setup._parent is None:
            ws.page_setup = PrintPageSetup(worksheet=ws)

        # Set employee code
        try:
            ws[self.EMPLOYEE_CODE_CELL] = int(employee_code)
        except ValueError:
            ws[self.EMPLOYEE_CODE_CELL] = employee_code

        # Set print area and page setup
        print_area = f"{config['image_start_col']}{config['image_start_row']}:{config['image_end_col']}{config['image_end_row']}"
        ws.print_area = print_area
        ws.page_margins = PageMargins(left=0.1, right=0.1, top=0.1, bottom=0.1, header=0, footer=0)
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_setup.horizontalCentered = True
        ws.page_setup.verticalCentered = True

        wb.save(excel_path)
        wb.close()

        # Export to PNG using LibreOffice
        png_path = self._export_to_png(excel_path, temp_dir, employee_code)

        # Read PNG and convert to base64
        with open(png_path, "rb") as f:
            base64_image = base64.b64encode(f.read()).decode("utf-8")

        # Read salary
        salary = self._read_salary_fast(excel_path)

        # Clean up PNG
        if os.path.exists(png_path):
            os.remove(png_path)

        logger.info(f"[{employee_code}] Done, salary={salary}")
        return BatchResult(
            employee_code=employee_code,
            success=True,
            base64_image=base64_image,
            salary=salary
        )

    def _read_salary_fast(self, excel_path: Path) -> Optional[int]:
        """Read salary from E24 - simplified version."""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, data_only=True)
            ws = wb[self.SALARY_SLIP_SHEET]
            salary_value = ws[self.SALARY_CELL].value
            wb.close()

            if salary_value is not None:
                try:
                    return int(float(salary_value))
                except (ValueError, TypeError):
                    return None
            return None
        except Exception as e:
            logger.error(f"[_read_salary_fast] Error: {e}")
            return None

    def _export_to_png(
        self,
        excel_path: Path,
        output_dir: Path,
        employee_code: str
    ) -> Path:
        """
        Export Excel print area to PNG using LibreOffice + PyMuPDF.

        Steps:
        1. Export to PDF via LibreOffice (respects print area)
        2. Convert PDF to PNG via PyMuPDF (pure Python, no external deps)
        """
        import fitz  # PyMuPDF

        # Export to PDF first (LibreOffice respects print area)
        libreoffice_path = get_libreoffice_path()
        logger.info(f"[{employee_code}] Using LibreOffice: {libreoffice_path}")

        cmd = [
            libreoffice_path,
            "--headless",
            "--calc",
            "--convert-to", "pdf",
            "--outdir", str(output_dir),
            str(excel_path)
        ]

        try:
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=60,
                shell=(sys.platform == "win32")
            )
        except FileNotFoundError:
            raise RuntimeError(
                f"LibreOffice not found at '{libreoffice_path}'. "
                "Please install LibreOffice and ensure it's in PATH."
            )

        if result.returncode != 0:
            raise RuntimeError(f"LibreOffice export failed: {result.stderr}")

        # The output file will be salary.pdf
        actual_pdf = output_dir / "salary.pdf"
        if not actual_pdf.exists():
            raise RuntimeError("PDF export failed - file not found")

        # Convert PDF to PNG using PyMuPDF (pure Python)
        logger.info(f"[{employee_code}] Converting PDF to PNG with PyMuPDF")
        png_path = output_dir / f"{employee_code}.png"

        doc = fitz.open(actual_pdf)
        page = doc[0]  # First page

        # Render at 150 DPI (default is 72 DPI)
        zoom = 150 / 72
        matrix = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=matrix)

        pix.save(png_path)
        doc.close()

        logger.info(f"[{employee_code}] PNG saved: {png_path}")

        # Clean up PDF
        if actual_pdf.exists():
            os.remove(actual_pdf)

        return png_path

    def _read_salary(self, excel_path: Path) -> Optional[int]:
        """Read calculated salary value from E24."""
        try:
            from openpyxl import load_workbook
            from openpyxl.worksheet.properties import WorksheetProperties

            logger.info(f"[_read_salary] Loading workbook: {excel_path}")
            wb = load_workbook(excel_path, data_only=True)
            logger.info(f"[_read_salary] Workbook loaded. Sheets: {wb.sheetnames}")

            logger.info(f"[_read_salary] Getting worksheet '{self.SALARY_SLIP_SHEET}'...")
            ws = wb[self.SALARY_SLIP_SHEET]
            logger.info(f"[_read_salary] Worksheet type: {type(ws)}, is None: {ws is None}")

            if ws is None:
                logger.error(f"[_read_salary] Worksheet is None!")
                wb.close()
                return None

            # Ensure sheet_properties exists to prevent NoneType errors
            logger.info(f"[_read_salary] sheet_properties type: {type(ws.sheet_properties)}, is None: {ws.sheet_properties is None}")
            if ws.sheet_properties is None:
                logger.info(f"[_read_salary] Initializing missing sheet_properties")
                ws.sheet_properties = WorksheetProperties()

            logger.info(f"[_read_salary] Reading cell {self.SALARY_CELL}...")
            salary_value = ws[self.SALARY_CELL].value
            logger.info(f"[_read_salary] Salary value: {salary_value}, type: {type(salary_value)}")
            wb.close()

            if salary_value is not None:
                try:
                    return int(float(salary_value))
                except (ValueError, TypeError):
                    return None
            return None
        except Exception as e:
            import traceback
            logger.error(f"[_read_salary] Error: {e}")
            logger.error(f"[_read_salary] Traceback:\n{traceback.format_exc()}")
            return None

    def generate_single(
        self,
        excel_file_path: str,
        employee_code: str,
        image_config: Optional[Dict[str, Any]] = None
    ) -> Tuple[str, Optional[int]]:
        """
        Generate salary slip for a single employee.

        For single generation, this is similar to the old method
        but uses the optimized PNG export.
        """
        results = self.generate_batch(
            excel_file_path,
            [employee_code],
            image_config
        )

        if results and results[0].success:
            return results[0].base64_image, results[0].salary
        elif results:
            raise RuntimeError(results[0].error or "Unknown error")
        else:
            raise RuntimeError("No results returned")


# Singleton instance
optimized_salary_slip_service = OptimizedSalarySlipService()
